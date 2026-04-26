"""
Sync Data.xlsx to Supabase DB.
Reads all competition and paper data from Excel, clears the DB, and re-inserts.
Poster images are extracted from Excel and uploaded to Supabase Storage.
Simplified Chinese text is automatically converted to Traditional Chinese.

Usage:
    python3 sync_db.py
"""

import requests
import json
import os
import zipfile
import openpyxl
from datetime import datetime
from opencc import OpenCC

# Simplified -> Traditional Chinese converter
_cc = OpenCC('s2t')

def s2t(text):
    """Convert Simplified Chinese to Traditional Chinese. Returns None if input is None."""
    if text is None:
        return None
    return _cc.convert(str(text).strip())

# ---- Config (reads from environment variables) ----
# Set these before running:
#   export SUPABASE_URL="https://xxx.supabase.co"
#   export SUPABASE_SERVICE_ROLE_KEY="eyJ..."
# Or put them in a .env file in this directory.

# Try loading from .env file
_env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
if os.path.exists(_env_path):
    with open(_env_path) as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip())

SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SERVICE_ROLE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")

if not SUPABASE_URL or not SERVICE_ROLE_KEY:
    print("❌ Missing SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY.")
    print("   Set them as env vars or create a .env file.")
    exit(1)

HEADERS = {
    "apikey": SERVICE_ROLE_KEY,
    "Authorization": f"Bearer {SERVICE_ROLE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(SCRIPT_DIR, "Data.xlsx")
POSTERS_DIR = os.path.join(SCRIPT_DIR, "posters")


def slugify(text):
    """Convert competition title to a URL-safe slug for poster filename."""
    import re
    slug = text.lower().strip()
    slug = re.sub(r'[^\w\s-]', '', slug)
    slug = re.sub(r'[\s_]+', '-', slug)
    slug = re.sub(r'-+', '-', slug)
    return slug[:60] + ".png"


def extract_and_upload_posters():
    """Extract poster images from Excel and upload to Supabase Storage."""
    print("📸 Extracting poster images from Excel...")

    # Ensure storage bucket exists
    resp = requests.post(
        f"{SUPABASE_URL}/storage/v1/bucket",
        headers={**HEADERS},
        json={"id": "posters", "name": "posters", "public": True},
    )

    # Extract images from Excel (xlsx is a ZIP file)
    images = []
    os.makedirs(POSTERS_DIR, exist_ok=True)

    with zipfile.ZipFile(EXCEL_PATH, "r") as z:
        for name in sorted(z.namelist()):
            if "media/image" in name:
                data = z.read(name)
                basename = os.path.basename(name)
                outpath = os.path.join(POSTERS_DIR, basename)
                with open(outpath, "wb") as f:
                    f.write(data)
                images.append(outpath)

    print(f"  Found {len(images)} images")
    return images


def upload_poster(local_path, slug):
    """Upload a single poster to Supabase Storage."""
    with open(local_path, "rb") as f:
        file_data = f.read()

    resp = requests.post(
        f"{SUPABASE_URL}/storage/v1/object/posters/{slug}",
        headers={
            "apikey": SERVICE_ROLE_KEY,
            "Authorization": f"Bearer {SERVICE_ROLE_KEY}",
            "Content-Type": "image/png",
            "x-upsert": "true",
        },
        data=file_data,
    )

    if resp.status_code in [200, 201]:
        return f"{SUPABASE_URL}/storage/v1/object/public/posters/{slug}"
    else:
        print(f"  ⚠️  Upload failed for {slug}: {resp.status_code}")
        return ""


def sync():
    """Main sync function: Excel -> Supabase."""
    print("🚀 HK UpUp - Syncing Data.xlsx to Supabase\n")

    # Step 1: Extract and upload posters
    image_paths = extract_and_upload_posters()

    # Step 2: Read Excel
    print("\n📖 Reading Excel data...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Competitions"]

    competitions = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row[0]:  # skip empty rows
            continue

        # Title: clean up newlines
        title = str(row[0]).replace("\n\n", " ").replace("\n", " ").strip()

        # Organizer: as-is
        organizer = str(row[1]).strip() if row[1] else ""

        # Deadline: datetime -> string
        deadline = row[2]
        if isinstance(deadline, datetime):
            deadline_str = deadline.strftime("%Y-%m-%d")
        else:
            deadline_str = str(deadline).strip() if deadline else None

        # Tags: split by ", " or ","
        raw_tags = str(row[3]).strip() if row[3] else ""
        tags = [t.strip() for t in raw_tags.replace(", ", ",").split(",") if t.strip()]

        # Prizes: as-is
        prizes = str(row[4]).strip() if row[4] else "TBC"

        # Official URL
        official_url = str(row[5]).strip() if row[5] else ""

        # Poster: upload and get URL
        poster_url = ""
        if i < len(image_paths):
            slug = slugify(title)
            print(f"  Uploading poster: {slug}")
            poster_url = upload_poster(image_paths[i], slug)

        # Status: auto-determine from deadline
        status = "open"
        if isinstance(row[2], datetime) and row[2] < datetime.now():
            status = "closed"

        competitions.append({
            "title": title,
            "organizer": organizer,
            "registration_deadline": deadline_str,
            "tags": tags,
            "prizes": prizes,
            "official_url": official_url,
            "poster_url": poster_url,
            "status": status,
        })

        print(f"  ✓ {title[:60]}")

    # Step 3: Clear DB
    print(f"\n🗑️  Clearing existing DB records...")
    requests.delete(
        f"{SUPABASE_URL}/rest/v1/competitions?title=neq.",
        headers=HEADERS,
    )

    # Step 4: Insert
    print(f"📥 Inserting {len(competitions)} competitions...")
    resp = requests.post(
        f"{SUPABASE_URL}/rest/v1/competitions",
        headers=HEADERS,
        json=competitions,
    )

    if resp.status_code in [200, 201]:
        data = resp.json()
        print(f"\n✅ Done! {len(data)} competitions synced to Supabase.")
    else:
        print(f"\n❌ Error: {resp.status_code} - {resp.text}")

    # Step 5: Sync Papers
    sync_papers(wb)


def sync_papers(wb):
    """Sync Papers sheet to Supabase papers table."""
    if "Papers" not in wb.sheetnames:
        print("\n⚠️  No 'Papers' sheet found in Excel, skipping.")
        return

    print("\n📄 Syncing Papers...")
    ws = wb["Papers"]

    papers = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # skip empty rows
            continue

        paper = {
            "full_name": s2t(row[0]),
            "abbreviation": s2t(row[1]) if row[1] else None,
            "official_url": str(row[2]).strip() if row[2] else None,
            "conference_dates": s2t(row[3]) if row[3] else None,
            "venue": s2t(row[4]) if row[4] else None,
            "organizer": s2t(row[5]) if row[5] else None,
            "registration_deadline": s2t(row[6]) if row[6] else None,
        }
        papers.append(paper)
        print(f"  ✓ {paper['abbreviation'] or paper['full_name'][:40]}")

    if not papers:
        print("  No paper data found.")
        return

    # Clear existing papers
    print(f"\n🗑️  Clearing existing papers...")
    requests.delete(
        f"{SUPABASE_URL}/rest/v1/papers?full_name=neq.",
        headers=HEADERS,
    )

    # Insert
    print(f"📥 Inserting {len(papers)} papers...")
    resp = requests.post(
        f"{SUPABASE_URL}/rest/v1/papers",
        headers=HEADERS,
        json=papers,
    )

    if resp.status_code in [200, 201]:
        data = resp.json()
        print(f"✅ {len(data)} papers synced to Supabase.")
    else:
        print(f"❌ Papers error: {resp.status_code} - {resp.text}")


if __name__ == "__main__":
    sync()
